import os
import pathlib
import zipfile
import csv
from PyPDF2 import PdfReader
from openpyxl import load_workbook

files_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')
resources_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'resources')
zip_path = os.path.join(resources_path, 'zip.zip')


# archived files
def test_archived_files():
    with zipfile.ZipFile(zip_path, 'w') as zip_file:
        for file in pathlib.Path(files_path).iterdir():
            zip_file.write(file, file.name)
    assert len(os.listdir(resources_path)) == 1

def test_read_pdf():
    with zipfile.ZipFile(zip_path) as p_f:
        pdf_archived = p_f.extract('pdf_file.pdf')
        reader = PdfReader(pdf_archived)
        assert len(reader.pages) == 412
        page = reader.pages[0]
        text = page.extract_text()
        assert 'Jul 14' in text
        os.remove('pdf_file.pdf')

def test_read_xlsx():
    with zipfile.ZipFile(zip_path) as x_f:
        xlsx_archived = x_f.extract('xlsx_file.xlsx')
        workbook = load_workbook(xlsx_archived)
        sheet = workbook.active
        assert sheet.cell(row=3, column=2).value == 'Mara'
        os.remove('xlsx_file.xlsx')


def test_read_csv():
    with zipfile.ZipFile(zip_path) as c_f:
        csv_archived = c_f.extract('csv_file.csv')
        with open(csv_archived) as csv_f:
            table = csv.reader(csv_f, delimiter=',')
            for index, line in enumerate(table):
                if index == 1:
                    assert 'Maria' in line
        os.remove('csv_file.csv')